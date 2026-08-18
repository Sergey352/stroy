#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Разбор Excel-файла сметы объекта в список позиций, готовых для загрузки
в справочник Pyrus «Смета».

СТРУКТУРА ИСХОДНОГО ФАЙЛА (проверено на реальном файле, сессия 2026-08-18):
  - Один файл = одна смета одного объекта.
  - Несколько листов в файле = категории 1-го уровня (инженерные системы:
    ВК1, ВК2, ОВ1, ОВ2 и т.п. — названия листов могут быть любыми).
  - Внутри листа встречаются строки, где объединены (merged) ячейки A:B
    и в них текст без числа в колонке «Поз» — это заголовки категории
    2-го уровня (подраздел сметы, например «Т3.2», «К2», «Электрическое
    оборудование»). Все строки-позиции ПОСЛЕ такого заголовка (и до
    следующего) относятся к этой категории 2-го уровня.
  - Строка-позиция (настоящий товар/материал) отличается от заголовков и
    промежуточных итогов тремя признаками одновременно:
      * колонка A («Поз») — число;
      * колонка C («Ед. изм.») — непустая строка;
      * колонка D («Кол-во») — число.
  - Имя объекта зашито в объединённую ячейку A4:D4 каждого листа, в виде
    "<Объект> -<КодЛиста>.СО" или "<Объект>-<КодЛиста>.СО" (пробел перед
    дефисом не всегда есть). Код листа совпадает с именем листа (например,
    лист "ВК1" -> суффикс "-ВК1.СО"), поэтому имя объекта извлекается
    отрезанием этого суффикса, а не регулярным выражением "по шаблону" —
    так надёжнее, потому что дефисы встречаются и внутри самого имени
    объекта (например "АКВАРЕЛЬ-Р").

КЛЮЧ УНИКАЛЬНОСТИ ПОЗИЦИИ (решено в сессии 2026-08-18):
  Наименование + Категория 2-го уровня. У одного и того же наименования
  в разных категориях 2-го уровня — это разные позиции сметы (например,
  один и тот же материал нужен и в одном подразделе, и в другом — суммарно
  их нужно больше). Если одинаковые Наименование+Категория встречаются в
  файле несколько раз (обычно — на разные точки монтажа), это настоящий
  дубль — количество суммируется в одну строку.
"""

import re
from collections import defaultdict

import openpyxl

# Колонки в исходном файле (см. заголовок строки 1 каждого листа):
# A=Поз, B=Наименование, C=Ед.изм., D=Кол-во, E=Цена за ед., F=Итого, G=Примечание
COL_POZ = 1
COL_NAME = 2
COL_UNIT = 3
COL_QTY = 4
COL_PRICE = 5
COL_TOTAL = 6
COL_NOTE = 7

# Данные (заголовки, промежуточные итоги) начинаются не раньше строки 4 —
# первые 3 строки заняты названием файла (A4 после слияния — тоже часть
# этой шапки) и двойным заголовком таблицы.
FIRST_DATA_ROW = 4


def _clean_text(value):
    """Приводит ячейку к чистой однострочной строке: убирает переносы
    строк (в исходном файле они встречаются примерно в 6% наименований —
    затрудняют чтение в списке справочника Pyrus) и лишние пробелы по
    краям. Возвращает None, если ячейка пустая."""
    if value is None:
        return None
    if not isinstance(value, str):
        return value
    text = value.replace("\n", " ").replace("\r", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text or None


def extract_object_name(ws, sheet_name):
    """Достаёт имя объекта из объединённой ячейки A4:D4 листа.

    В файле заголовок листа выглядит как "<Объект> -<ИмяЛиста>.СО" или
    "<Объект>-<ИмяЛиста>.СО". Мы точно знаем ИмяЛиста (это и есть имя
    текущего листа), поэтому просто отрезаем всё начиная с первого
    вхождения "-<ИмяЛиста>." — это надёжнее, чем угадывать по регулярке,
    потому что в самом имени объекта тоже есть дефисы.
    """
    raw = ws.cell(row=4, column=1).value
    if not isinstance(raw, str):
        return None
    marker = f"-{sheet_name}."
    idx = raw.find(marker)
    if idx == -1:
        # Не нашли ожидаемый суффикс — возвращаем текст как есть (без
        # обрезки), чтобы не потерять информацию молча.
        return _clean_text(raw)
    return _clean_text(raw[:idx])


def parse_smeta_file(path):
    """Читает xlsx-файл сметы и возвращает список «сырых» позиций (ещё
    без агрегации одинаковых Наименование+Категория) в виде словарей:
        {
            "object": ...,   # имя объекта, из ячейки A4 листа
            "cat1": ...,     # категория 1-го уровня = имя листа
            "cat2": ...,     # категория 2-го уровня = ближайший заголовок-группа выше
            "name": ...,     # наименование позиции
            "unit": ...,     # единица измерения
            "qty": ...,      # количество по смете
            "price": ...,    # цена за единицу (может быть None)
        }
    Обрабатывает КАЖДЫЙ лист файла отдельно (категория 1-го уровня = лист).
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    positions = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        object_name = extract_object_name(ws, sheet_name)

        # Собираем список строк, которые являются объединёнными
        # заголовками-группами (merged-диапазон в пределах одной строки).
        # Именно такие строки и есть заголовки категории 2-го уровня —
        # НО только если в колонке «Поз» при этом нет числа (иначе это
        # просто отформатированная позиция, а не заголовок).
        merged_row_labels = {}
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_row == merged_range.max_row:
                merged_row_labels[merged_range.min_row] = ws.cell(
                    row=merged_range.min_row, column=merged_range.min_col
                ).value

        current_cat2 = None
        for row in range(FIRST_DATA_ROW, ws.max_row + 1):
            poz = ws.cell(row=row, column=COL_POZ).value
            name = ws.cell(row=row, column=COL_NAME).value
            unit = ws.cell(row=row, column=COL_UNIT).value
            qty = ws.cell(row=row, column=COL_QTY).value
            price = ws.cell(row=row, column=COL_PRICE).value

            is_group_header = row in merged_row_labels and not isinstance(poz, (int, float))
            if is_group_header:
                current_cat2 = _clean_text(merged_row_labels[row])
                continue

            is_position = (
                isinstance(poz, (int, float))
                and isinstance(unit, str)
                and unit.strip() != ""
                and isinstance(qty, (int, float))
            )
            if not is_position:
                continue  # промежуточный итог / пустая строка / прочий мусор

            positions.append(
                {
                    "object": object_name,
                    "cat1": sheet_name,
                    "cat2": current_cat2,
                    "name": _clean_text(name),
                    "unit": _clean_text(unit),
                    "qty": float(qty),
                    "price": float(price) if isinstance(price, (int, float)) else None,
                }
            )

    return positions


def aggregate_positions(positions):
    """Схлопывает «сырые» позиции по ключу (Объект, Наименование,
    Категория 2-го уровня): количество суммируется, цена берётся из
    первой встреченной строки (в пределах одного наименования+категории
    цена почти всегда одинаковая — построчных расхождений в проверенном
    файле не было).

    Возвращает список агрегированных позиций той же формы, что и на
    входе, но уже без дублей внутри одного (объект, наименование, cat2).
    """
    groups = defaultdict(list)
    for pos in positions:
        key = (pos["object"], pos["name"], pos["cat2"])
        groups[key].append(pos)

    aggregated = []
    for (object_name, name, cat2), items in groups.items():
        aggregated.append(
            {
                "object": object_name,
                "cat1": items[0]["cat1"],
                "cat2": cat2,
                "name": name,
                "unit": items[0]["unit"],
                "qty": sum(i["qty"] for i in items),
                "price": items[0]["price"],
            }
        )
    return aggregated


if __name__ == "__main__":
    # Небольшая самопроверка при прямом запуске файла: разобрать пример
    # и напечатать сводку — удобно для быстрой проверки без отдельного
    # скрипта импорта.
    import sys

    path = sys.argv[1] if len(sys.argv) > 1 else (
        "ДЛЯ_ПАЙРУСА_КП_АКВАРЕЛЬ_ОВВК_29_06_2026_—_копия.xlsx"
    )
    raw = parse_smeta_file(path)
    agg = aggregate_positions(raw)
    print(f"Сырых позиций: {len(raw)}")
    print(f"После агрегации (объект+имя+категория): {len(agg)}")
    objects = sorted(set(p["object"] for p in agg))
    print(f"Объекты в файле: {objects}")
